"""
HFC Inputs Auto-Populator — Web App
=====================================
Run with:  streamlit run app.py
Share with your team by sending them:  http://<your-ip-address>:8501

Install dependencies first:
    pip install streamlit openpyxl google-api-python-client google-auth-httplib2 google-auth-oauthlib boxsdk
"""

import io
import os
import re
import tempfile
import urllib.parse

import streamlit as st

# ─────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="HFC Inputs Auto-Populator",
    page_icon="📋",
    layout="centered",
)

st.title("📋 HFC Inputs Auto-Populator")
st.caption("Innovations for Poverty Action · Data Management Tool")
st.markdown("---")
st.markdown(
    "Upload your SurveyCTO XLSForm or pull it directly from Google Drive — "
    "the tool will automatically populate all six DMS input sheets "
    "*(other specify, outliers, constraints, logic, enumstats, text audits)*. "
    "The HFC inputs template is pre-loaded, and the populated file can be saved "
    "to your Box folder or downloaded directly from your browser."
)
st.caption("Built by Jimmy Jairo · Innovations for Poverty Action Tanzania")

# ─────────────────────────────────────────────────────────────
# Core population logic (shared by both tabs)
# ─────────────────────────────────────────────────────────────

def clean_label(s):
    if not s:
        return ""
    return re.sub(r"<[^>]+>", "", str(s)).strip()


# SurveyCTO surveys frequently use these as missing/refused/don't-know codes.
# Clauses comparing the field to one of these are excluded from bound parsing.
MISSING_CODES = {-888, -999, -666, -777, -88, -99}


def parse_constraint_bounds(expr):
    """Extract (hard_min, hard_max) from a SurveyCTO constraint expression.

    Handles common patterns like:
        .>0 and .<100              -> (0, 100)
        .>=0 and .<10              -> (0, 10)
        .>= 0 or .= -888 or .= -999 -> (0, None)
        . between 0 and 24          -> (0, 24)

    Branches comparing the field to a missing-value code are dropped.
    Clauses referencing other variables (${var}) are dropped — those are
    cross-checks, not magnitude bounds. Returns (None, None) if nothing
    parseable was found.
    """
    if not expr:
        return (None, None)
    s = str(expr).strip().strip("()").strip()
    if not s:
        return (None, None)

    # ". between N and M" -> handle before the AND-split would shred it
    m = re.match(r"^\s*\.\s*between\s+(-?\d+(?:\.\d+)?)\s+and\s+(-?\d+(?:\.\d+)?)\s*$",
                 s, flags=re.IGNORECASE)
    if m:
        return (_to_int_if_whole(float(m.group(1))),
                _to_int_if_whole(float(m.group(2))))

    # Disjunctive branches: "<bound clauses> or .= -888 or .= -999"
    branches = re.split(r"\s+or\s+", s, flags=re.IGNORECASE)
    real_branches = []
    for br in branches:
        br_clean = br.strip().strip("()").strip()
        # Drop pure missing-code equality branches
        mc = re.match(r"^\s*\.\s*=\s*(-?\d+(?:\.\d+)?)\s*$", br_clean)
        if mc and float(mc.group(1)) in MISSING_CODES:
            continue
        # Drop branches that reference other variables
        if "${" in br_clean:
            continue
        real_branches.append(br_clean)

    if not real_branches:
        return (None, None)

    # Multiple disjunctive non-missing branches can't be safely combined into
    # a single magnitude bound, so take the first (typical surveys have one
    # bound branch plus missing-code branches).
    branch = real_branches[0]

    mins, maxs = [], []
    for cl in re.split(r"\s+and\s+", branch, flags=re.IGNORECASE):
        cl = cl.strip().strip("()").strip()
        if not cl or "${" in cl:
            continue
        m = re.match(r"^\s*\.\s*(>=|<=|>|<|=)\s*(-?\d+(?:\.\d+)?)\s*$", cl)
        if not m:
            continue
        op, val = m.group(1), float(m.group(2))
        if op == "=":
            continue
        if op in (">=", ">"):
            mins.append(val)
        else:
            maxs.append(val)

    hard_min = max(mins) if mins else None
    hard_max = min(maxs) if maxs else None
    return (_to_int_if_whole(hard_min), _to_int_if_whole(hard_max))


def _to_int_if_whole(v):
    if v is None:
        return None
    return int(v) if v == int(v) else v


def classify_survey_variables(survey_bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(survey_bytes), read_only=True, data_only=True)
    ws = wb["survey"]

    SKIP_TYPES = {
        "begin group", "end group", "begin repeat", "end repeat",
        "note", "calculate", "start", "end", "deviceid",
        "phonenumber", "username", "caseid", "enumerator",
    }

    # Build a header-name -> column-index map so we work with any XLSForm
    # column order (SurveyCTO/ODK templates vary; e.g. some forms include
    # `accuracy_threshold` and others don't, shifting later columns).
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, ()) or ()
    col_idx = {str(h).strip().lower(): i for i, h in enumerate(header_row) if h}

    def cell(row, name, default=""):
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        return val if val is not None else default

    groups, numerics, selects, texts = [], [], [], []
    repeat_depth = 0

    for row in rows_iter:
        t = cell(row, "type")
        n = cell(row, "name")
        l = clean_label(cell(row, "label"))
        constraint = cell(row, "constraint", default=None)
        disabled = str(cell(row, "disabled")).strip().lower() == "yes"
        t_str = str(t).lower() if t else ""

        # Track repeat-group nesting depth (skip disabled begin/end repeat
        # so depth stays balanced for them too)
        if "begin" in t_str and "repeat" in t_str:
            if not disabled:
                repeat_depth += 1
            continue
        if "end" in t_str and "repeat" in t_str:
            if not disabled:
                repeat_depth = max(0, repeat_depth - 1)
            continue

        if disabled:
            continue
        if not n:
            continue

        in_repeat = repeat_depth > 0
        base = t.split()[0] if t else ""
        if "begin" in str(t) and "group" in str(t):
            groups.append({"name": n, "label": l, "in_repeat": in_repeat})
        elif base in SKIP_TYPES or not base:
            continue
        elif base in ("integer", "decimal"):
            numerics.append({"name": n, "type": t, "label": l,
                             "in_repeat": in_repeat, "constraint": constraint})
        elif base in ("select_one", "select_multiple"):
            selects.append({"name": n, "type": t, "label": l, "in_repeat": in_repeat})
        elif base == "text":
            texts.append({"name": n, "label": l, "in_repeat": in_repeat})

    wb.close()
    return numerics, selects, texts, groups


def fmt_name(v):
    """Append * to variable name when it lives inside a repeat group, so
    Stata can match the wide-format instances (e.g. var_r1, var_r2 -> var_r*)."""
    name = str(v["name"]).strip()
    return f"{name}*" if v.get("in_repeat") else name


def existing_vars(ws, col=0):
    return {str(row[col]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[col]}


_OTHER_SUFFIXES = ("_oth", "_other", "_specify", "_oth_r", "_other_r", "_specify_r")


def populate_other_specify(ws, selects, texts):
    text_by_name = {v["name"]: v for v in texts}
    existing = existing_vars(ws)
    added = 0
    for v in selects:
        name = v["name"]
        for suffix in _OTHER_SUFFIXES:
            child = name + suffix
            child_var = text_by_name.get(child)
            if child_var and name not in existing:
                child_label = child_var.get("label") or "Other, specify"
                ws.append((fmt_name(v), v["label"], fmt_name(child_var), child_label,
                           None, None, "id member_name village enumerator_id"))
                existing.add(name)
                added += 1
                break
    return added


def populate_outliers(ws, numerics):
    existing = existing_vars(ws)
    added = 0
    for v in numerics:
        if v["name"] not in existing:
            combine = "yes" if v.get("in_repeat") else None
            ws.append((fmt_name(v), v["label"], "enum", "sd", 3, combine, None, None, "id member_name enumerator_id"))
            added += 1
    return added


def populate_constraints(ws, numerics):
    """Hard min/max are pulled directly from each numeric's SurveyCTO
    `constraint` expression. Soft bounds are intentionally left blank for
    the user to calibrate per study — there's no defensible way to invent
    them from the form alone."""
    existing = existing_vars(ws)
    added = 0
    for v in numerics:
        if v["name"] in existing:
            continue
        hard_min, hard_max = parse_constraint_bounds(v.get("constraint"))
        ws.append((fmt_name(v), v["label"], hard_min, None, None, hard_max, None, None, None))
        added += 1
    return added


def populate_logic(ws, numerics):
    """Logic checks are study-specific cross-variable consistency rules and
    cannot reliably be derived from XLSForm metadata alone — the constraint
    column rarely encodes them, and patterns like annual/monthly pairs are
    naming-convention-dependent. Sheet is left for the user to fill in."""
    return 0


def populate_enumstats(ws, numerics):
    existing = existing_vars(ws)
    added = 0
    for v in numerics:
        if v["name"] not in existing:
            combine = "yes" if v.get("in_repeat") else None
            ws.append((fmt_name(v), v["label"], "yes", None, "number", "yes", "yes", "yes", combine, None))
            added += 1
    return added


def populate_text_audits(ws, groups):
    """Emit every top-level begin-group from the form. The user can prune
    the ones they don't want timed in text audits."""
    existing = existing_vars(ws)
    added = 0
    for g in groups:
        if g["name"] not in existing:
            ws.append((fmt_name(g), None, None, g["label"]))
            added += 1
    return added


def clear_data_rows(ws):
    """Delete all rows below the header (row 1) so population always starts at row 2."""
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)


def run_population(survey_bytes: bytes, template_bytes: bytes) -> bytes:
    from openpyxl import load_workbook

    numerics, selects, texts, groups = classify_survey_variables(survey_bytes)

    wb = load_workbook(io.BytesIO(template_bytes), keep_vba=True)

    # Clear existing data rows so output always starts at row 2
    for sheet_name in ["other specify", "outliers", "constraints", "logic", "enumstats", "text audits"]:
        clear_data_rows(wb[sheet_name])

    results = {
        "other specify": populate_other_specify(wb["other specify"], selects, texts),
        "outliers":      populate_outliers(wb["outliers"], numerics),
        "constraints":   populate_constraints(wb["constraints"], numerics),
        "logic":         populate_logic(wb["logic"], numerics),
        "enumstats":     populate_enumstats(wb["enumstats"], numerics),
        "text audits":   populate_text_audits(wb["text audits"], groups),
    }

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), results, len(numerics), len(selects), len(texts)


def show_results(output_bytes, results, n_num, n_sel, n_txt, filename):
    st.success("Populated successfully!")

    col1, col2, col3 = st.columns(3)
    col1.metric("Numeric variables", n_num)
    col2.metric("Select variables", n_sel)
    col3.metric("Text variables", n_txt)

    st.markdown("**Rows added per sheet:**")
    for sheet, count in results.items():
        icon = "✅" if count > 0 else "➖"
        st.markdown(f"{icon} `{sheet}` — **+{count}** rows")

    st.download_button(
        label="⬇️  Download populated HFC inputs file",
        data=output_bytes,
        file_name=filename,
        mime="application/vnd.ms-excel.sheet.macroEnabled.12",
        use_container_width=True,
    )


# ─────────────────────────────────────────────────────────────
# Google OAuth helpers
# ─────────────────────────────────────────────────────────────

def _oauth_configured() -> bool:
    """Return True if Google OAuth client credentials are in Streamlit secrets."""
    try:
        return bool(st.secrets.get("GOOGLE_CLIENT_ID") and st.secrets.get("GOOGLE_CLIENT_SECRET"))
    except Exception:
        return False

def _oauth_redirect_uri() -> str:
    try:
        return st.secrets.get("GOOGLE_REDIRECT_URI", "http://localhost:8501")
    except Exception:
        return "http://localhost:8501"

def _build_oauth_url() -> str:
    params = {
        "client_id": st.secrets["GOOGLE_CLIENT_ID"],
        "redirect_uri": _oauth_redirect_uri(),
        "response_type": "code",
        "scope": "https://www.googleapis.com/auth/drive.readonly openid email profile",
        "access_type": "offline",
        "prompt": "select_account",
    }
    return "https://accounts.google.com/o/oauth2/v2/auth?" + urllib.parse.urlencode(params)

def _exchange_oauth_code(code: str) -> dict:
    import requests
    r = requests.post("https://oauth2.googleapis.com/token", data={
        "code": code,
        "client_id": st.secrets["GOOGLE_CLIENT_ID"],
        "client_secret": st.secrets["GOOGLE_CLIENT_SECRET"],
        "redirect_uri": _oauth_redirect_uri(),
        "grant_type": "authorization_code",
    }, timeout=15)
    return r.json()

def _fetch_google_user(access_token: str) -> dict:
    import requests
    r = requests.get(
        "https://www.googleapis.com/oauth2/v2/userinfo",
        headers={"Authorization": f"Bearer {access_token}"},
        timeout=10,
    )
    return r.json() if r.status_code == 200 else {}

def _download_authed(file_id: str, access_token: str) -> bytes:
    """Download a Drive file using an authenticated access token."""
    import requests
    h = {"Authorization": f"Bearer {access_token}"}

    # Try export as xlsx (works for native Google Sheets)
    r = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}/export"
        "?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=h, timeout=30,
    )
    if r.status_code == 200 and r.content[:2] == b"PK":
        return r.content

    # Direct download (uploaded .xlsx / XLSForm files)
    r = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media",
        headers=h, timeout=60,
    )
    if r.status_code == 200 and r.content[:2] == b"PK":
        return r.content

    try:
        err_detail = r.json().get("error", {}).get("message", r.text[:200])
    except Exception:
        err_detail = r.text[:200]
    raise ValueError(
        f"Authenticated download failed (HTTP {r.status_code}): {err_detail}\n\n"
        "Common causes:\n"
        "• Google Drive API is not enabled in your Google Cloud project\n"
        "• The file has not been shared with your Google account"
    )


# ─────────────────────────────────────────────────────────────
# Handle Google OAuth callback (runs before tabs render)
# ─────────────────────────────────────────────────────────────
_qp = st.query_params.to_dict()
if "code" in _qp and "google_token" not in st.session_state:
    with st.spinner("Connecting to Google Drive..."):
        try:
            _tok = _exchange_oauth_code(_qp["code"])
            if "access_token" in _tok:
                st.session_state["google_token"] = _tok
                st.session_state["google_user"] = _fetch_google_user(_tok["access_token"])
            else:
                st.error(
                    f"Google sign-in failed: "
                    f"{_tok.get('error_description', _tok.get('error', 'Unknown error'))}"
                )
        except Exception as _e:
            st.error(f"Google sign-in error: {_e}")
        finally:
            st.query_params.clear()
            st.rerun()


# ─────────────────────────────────────────────────────────────
# Tab 1: Upload files directly
# Tab 2: Use Google Drive + Box IDs
# ─────────────────────────────────────────────────────────────

tab1, tab2 = st.tabs(["📁  Upload Files", "☁️  Google Drive & Box"])

# ── TAB 1: UPLOAD ────────────────────────────────────────────
with tab1:
    st.subheader("Upload your survey form")
    st.markdown("Upload your SurveyCTO XLSForm. The HFC inputs template is pre-loaded.")

    survey_file = st.file_uploader("SurveyCTO XLSForm (.xlsx)", type=["xlsx"], key="upload_survey")

    # Load bundled template
    BUNDLED_TEMPLATE = os.path.join(os.path.dirname(__file__), "hfc_inputs.xlsm")
    has_bundled = os.path.exists(BUNDLED_TEMPLATE)

    if has_bundled:
        st.success("HFC inputs template: `hfc_inputs.xlsm` (pre-loaded)")
    else:
        st.warning("No bundled template found — please upload one below.")

    template_file = None
    if not has_bundled:
        template_file = st.file_uploader("HFC Inputs Template (.xlsm)", type=["xlsm"], key="upload_template")

    ready = survey_file and (has_bundled or template_file)

    if ready:
        out_name = "hfc_inputs_populated.xlsm"
        if st.button("Run Auto-Populate", key="btn_upload", use_container_width=True, type="primary"):
            with st.spinner("Populating HFC inputs..."):
                try:
                    survey_bytes = survey_file.read()
                    if has_bundled:
                        with open(BUNDLED_TEMPLATE, "rb") as f:
                            template_bytes = f.read()
                    else:
                        template_bytes = template_file.read()
                    output_bytes, results, n_num, n_sel, n_txt = run_population(
                        survey_bytes, template_bytes
                    )
                    show_results(output_bytes, results, n_num, n_sel, n_txt, out_name)
                except Exception as e:
                    st.error(f"Something went wrong: {e}")
    else:
        st.info("Upload the survey form above to enable the run button.")


# ── TAB 2: GOOGLE DRIVE LINK + BOX PATH ──────────────────────
with tab2:
    st.subheader("Google Drive link & Box folder path")
    st.markdown(
        "Sign in with your Google work email to access survey forms shared with you, "
        "then paste the file link and choose where to save the output."
    )

    # ── 0. Google account sign-in ──────────────────────────────
    st.markdown("**0. Google account**")
    _google_configured = _oauth_configured()
    _google_token = st.session_state.get("google_token")
    _google_user  = st.session_state.get("google_user", {})

    if not _google_configured:
        st.info(
            "Google sign-in is not set up yet. "
            "Ask your admin to add `GOOGLE_CLIENT_ID`, `GOOGLE_CLIENT_SECRET`, and "
            "`GOOGLE_REDIRECT_URI` to the app's Streamlit secrets. "
            "Files shared publicly will still work without sign-in."
        )
    elif _google_token:
        _email = _google_user.get("email", "your account")
        col_a, col_b = st.columns([7, 2])
        col_a.success(f"Signed in as **{_email}**")
        if col_b.button("Sign out", key="google_signout"):
            st.session_state.pop("google_token", None)
            st.session_state.pop("google_user", None)
            st.rerun()
    else:
        st.warning("Sign in with your Google work email to access files shared with you.")
        st.link_button(
            "Sign in with Google",
            _build_oauth_url(),
            use_container_width=True,
        )

    st.markdown("")

    # ── Helper: extract file ID from Google Drive URL ──────────
    def extract_gdrive_id(url: str):
        import re
        patterns = [
            r"/d/([a-zA-Z0-9_-]{20,})",
            r"id=([a-zA-Z0-9_-]{20,})",
        ]
        for p in patterns:
            m = re.search(p, url)
            if m:
                return m.group(1)
        return None

    def download_gdrive_file(file_id: str) -> bytes:
        import requests, re as _re

        def is_excel(data: bytes) -> bool:
            return len(data) > 4 and data[:2] == b"PK"

        # ── Authenticated download (if user is signed in) ──────
        _tok = st.session_state.get("google_token")
        if _tok:
            # Signed in — use auth only, surface the real error if it fails
            return _download_authed(file_id, _tok["access_token"])

        # ── Public download fallback ───────────────────────────
        session = requests.Session()
        headers = {"User-Agent": "Mozilla/5.0"}

        # Attempt 1: Direct download with confirm=t (uploaded .xlsx / XLSForm)
        r = session.get(
            f"https://drive.google.com/uc?export=download&confirm=t&id={file_id}",
            timeout=60, headers=headers,
        )
        if r.status_code == 200 and is_excel(r.content):
            return r.content

        # Attempt 2: Spreadsheet export (native Google Sheets)
        r = session.get(
            f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx",
            timeout=30, headers=headers,
        )
        if r.status_code == 200 and is_excel(r.content):
            return r.content

        # Attempt 3: Follow confirmation token for large files
        r = session.get(
            f"https://drive.google.com/uc?export=download&id={file_id}",
            timeout=30, headers=headers,
        )
        if not is_excel(r.content):
            token_match = _re.search(rb'confirm=([^&"]+)', r.content)
            if token_match:
                r = session.get(
                    f"https://drive.google.com/uc?export=download"
                    f"&confirm={token_match.group(1).decode()}&id={file_id}",
                    timeout=60, headers=headers,
                )
        if r.status_code == 200 and is_excel(r.content):
            return r.content

        raise ValueError(
            "Could not download the file. Please:\n"
            "• Sign in with Google (button above) to access files shared with your work email, OR\n"
            "• Set the file's sharing to **'Anyone with the link can view'** for public access."
        )

    # ── Helper: clean and resolve any local path ──────────────
    def resolve_path(raw: str):
        from pathlib import Path
        # Only strip genuinely invisible/null characters, nothing else
        clean = raw
        for bad in ["\x00", "​", "‌", "‍", "﻿", " "]:
            clean = clean.replace(bad, "")
        clean = clean.strip().strip('"').strip("'").strip()
        return Path(clean), clean

    # ── 1. Google Drive survey link (now after sign-in section) ─
    st.markdown("**1. Survey form — Google Drive share link**")
    gdrive_url = st.text_input(
        "Paste the Google Drive share link",
        placeholder="https://docs.google.com/spreadsheets/d/.../edit?usp=sharing",
        label_visibility="collapsed",
    )
    gdrive_file_id = None
    if gdrive_url:
        gdrive_file_id = extract_gdrive_id(gdrive_url)
        if gdrive_file_id:
            st.success(f"File ID detected: `{gdrive_file_id}`")
        else:
            st.error("Could not read a file ID from that URL. Make sure you copied the full share link.")

    # ── 2. HFC inputs template ─────────────────────────────────
    BUNDLED_TEMPLATE = os.path.join(os.path.dirname(__file__), "hfc_inputs.xlsm")
    has_bundled = os.path.exists(BUNDLED_TEMPLATE)

    st.markdown("**2. HFC inputs template**")
    template_source = st.radio(
        "Template source",
        options=["Use pre-loaded template", "Use template from folder path"],
        label_visibility="collapsed",
        horizontal=True,
    )

    template_bytes_cloud = None
    if template_source == "Use pre-loaded template":
        if has_bundled:
            st.success("Using pre-loaded `hfc_inputs.xlsm`")
            with open(BUNDLED_TEMPLATE, "rb") as f:
                template_bytes_cloud = f.read()
        else:
            st.warning("No pre-loaded template found. Switch to folder path option below.")
    else:
        template_path_input = st.text_input(
            "Paste the full file path to your HFC template (.xlsm)",
            placeholder=r"C:\Users\yourname\Box\...\hfc_inputs.xlsm",
            key="template_path_input",
        )
        if template_path_input:
            tp, tp_clean = resolve_path(template_path_input)
            st.caption(f"Looking for: `{tp_clean}`")
            if tp.is_file():
                st.success(f"Template found: `{tp.name}`")
                template_bytes_cloud = tp.read_bytes()
            elif tp.is_dir():
                xlsm_files = sorted(tp.glob("*.xlsm"))
                if xlsm_files:
                    chosen = st.selectbox("Select template", xlsm_files, format_func=lambda f: f.name)
                    template_bytes_cloud = chosen.read_bytes()
                else:
                    st.error("No .xlsm files found in that folder.")
            else:
                st.error("Path not found. Check the path is correct and the file exists.")

    # ── 3. Output location ─────────────────────────────────────
    import platform
    is_windows_local = platform.system() == "Windows"

    output_name = "hfc_inputs_populated.xlsm"
    out_folder = None

    if is_windows_local:
        st.markdown("**3. Where to save the output — folder path**")
        output_path_input = st.text_input(
            "Paste a folder path or full file path for the output",
            placeholder=r"C:\Users\yourname\Desktop\DMS App\3_checks\1_inputs",
            label_visibility="collapsed",
            key="output_path_input",
        )
        if output_path_input:
            op, op_clean = resolve_path(output_path_input)
            if op_clean.lower().endswith(".xlsm"):
                out_folder = op.parent
                output_name = op.name
            else:
                out_folder = op
            if out_folder and out_folder.exists():
                st.success(f"Output folder found: `{out_folder}`")
            else:
                st.error(
                    "Folder not found. Check the path is correct.\n\n"
                    "**Tip:** Open the folder in Windows Explorer → click the address bar → copy the path."
                )
                out_folder = None
        output_name = st.text_input("Output filename", value=output_name, key="output_name_cloud")
    else:
        # Running on Streamlit Cloud — just download via browser
        st.markdown("**3. Output**")
        st.info("The populated file will be ready to download directly in your browser.")
        out_folder = "download_only"  # sentinel value

    # ── Run button ─────────────────────────────────────────────
    st.markdown("")
    ready = gdrive_file_id and template_bytes_cloud and out_folder

    if st.button("Run Auto-Populate", key="btn_cloud", use_container_width=True, type="primary", disabled=not ready):
        with st.spinner("Downloading survey and populating..."):
            try:
                survey_bytes = download_gdrive_file(gdrive_file_id)
                output_bytes, results, n_num, n_sel, n_txt = run_population(
                    survey_bytes, template_bytes_cloud
                )
                if is_windows_local and out_folder != "download_only":
                    from pathlib import Path
                    out_path = Path(str(out_folder)) / output_name
                    out_path.write_bytes(output_bytes)
                    show_results(output_bytes, results, n_num, n_sel, n_txt, output_name)
                    st.info(f"Saved to: `{out_path}`")
                else:
                    show_results(output_bytes, results, n_num, n_sel, n_txt, output_name)
            except Exception as e:
                st.error(f"Something went wrong: {e}")

    if not ready:
        missing = []
        if not gdrive_file_id: missing.append("Google Drive survey link")
        if not template_bytes_cloud: missing.append("HFC inputs template")
        if is_windows_local and not out_folder: missing.append("output folder path")
        if missing:
            st.info(f"Still needed: {', '.join(missing)}")


# ─────────────────────────────────────────────────────────────
# Footer
# ─────────────────────────────────────────────────────────────
st.markdown("---")
st.caption("Built by Jimmy Jairo · IPA Tanzania · jjairo@poverty-action.org")
