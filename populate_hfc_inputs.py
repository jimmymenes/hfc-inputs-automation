"""
HFC Inputs Auto-Populator
=========================
Pulls a SurveyCTO XLSForm from Google Drive and an HFC inputs template
from Box, populates all six DMS sheets (other specify, outliers,
constraints, logic, enumstats, text audits), and uploads the result
back to Box.

SETUP — run once before first use:
    pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib boxsdk openpyxl

HOW TO USE:
    1. Fill in the CONFIG section below with your file IDs and credentials.
    2. Run:  python populate_hfc_inputs.py
    3. The populated file will appear in the same Box folder as your template.

GETTING FILE IDs:
    Google Drive: open the file → copy the long ID from the URL
        https://drive.google.com/file/d/<FILE_ID>/view
    Box: open the file → copy the number at the end of the URL
        https://app.box.com/file/<FILE_ID>

CREDENTIALS:
    Google Drive: download credentials.json from Google Cloud Console
        (APIs & Services → Credentials → OAuth 2.0 Client → Download)
    Box: create an app at https://developer.box.com
        (Custom App → User Authentication OAuth 2.0 → copy Client ID & Secret)
"""

import io
import os
import re
import tempfile

# ============================================================
# CONFIG — edit these values
# ============================================================

# Google Drive: file ID of your XLSForm survey (.xlsx)
GDRIVE_SURVEY_FILE_ID = "YOUR_GOOGLE_DRIVE_FILE_ID_HERE"

# Box: file ID of your HFC inputs template (.xlsm)
BOX_TEMPLATE_FILE_ID = "YOUR_BOX_TEMPLATE_FILE_ID_HERE"

# Box: folder ID where the populated output file should be saved
# (find it in the Box URL when you open the folder)
BOX_OUTPUT_FOLDER_ID = "YOUR_BOX_OUTPUT_FOLDER_ID_HERE"

# Output filename saved to Box
OUTPUT_FILENAME = "hfc_inputs_populated.xlsm"

# Path to your Google OAuth credentials JSON file
GOOGLE_CREDENTIALS_FILE = "google_credentials.json"

# Box credentials
BOX_CLIENT_ID     = "YOUR_BOX_CLIENT_ID"
BOX_CLIENT_SECRET = "YOUR_BOX_CLIENT_SECRET"

# ============================================================
# STEP 1: Authenticate and download survey from Google Drive
# ============================================================

def get_google_drive_service():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build

    SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
    token_file = "google_token.json"
    creds = None

    if os.path.exists(token_file):
        creds = Credentials.from_authorized_user_file(token_file, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(GOOGLE_CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_file, "w") as f:
            f.write(creds.to_json())

    return build("drive", "v3", credentials=creds)


def download_from_google_drive(file_id: str, dest_path: str):
    from googleapiclient.http import MediaIoBaseDownload

    service = get_google_drive_service()
    request = service.files().get_media(fileId=file_id)
    with open(dest_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    print(f"  Downloaded survey from Google Drive → {dest_path}")


# ============================================================
# STEP 2: Authenticate and download template from Box
# ============================================================

def get_box_client():
    from boxsdk import OAuth2, Client

    auth = OAuth2(
        client_id=BOX_CLIENT_ID,
        client_secret=BOX_CLIENT_SECRET,
        store_tokens=_store_box_tokens,
        access_token=_load_box_token("access"),
        refresh_token=_load_box_token("refresh"),
    )
    return Client(auth)


def _store_box_tokens(access_token, refresh_token):
    with open("box_token.txt", "w") as f:
        f.write(f"{access_token}\n{refresh_token}")


def _load_box_token(kind: str):
    if not os.path.exists("box_token.txt"):
        return None
    with open("box_token.txt") as f:
        lines = f.read().splitlines()
    return lines[0] if kind == "access" and lines else (lines[1] if kind == "refresh" and len(lines) > 1 else None)


def box_first_time_auth():
    """
    Run this once to get your first Box tokens.
    Opens a browser window for you to log in.
    """
    from boxsdk import OAuth2, Client

    auth_url_visited = {}

    def _open_browser(url):
        import webbrowser
        print(f"\n  Opening Box login in your browser...")
        webbrowser.open(url)

    oauth = OAuth2(
        client_id=BOX_CLIENT_ID,
        client_secret=BOX_CLIENT_SECRET,
        store_tokens=_store_box_tokens,
    )
    auth_url, csrf_token = oauth.get_authorization_url("http://localhost")
    _open_browser(auth_url)
    redirect_url = input("  Paste the full redirect URL from your browser here: ").strip()
    code = redirect_url.split("code=")[1].split("&")[0]
    access_token, refresh_token = oauth.authenticate(code)
    print("  Box authentication successful. Tokens saved to box_token.txt")
    return Client(oauth)


def download_from_box(file_id: str, dest_path: str):
    client = get_box_client()
    box_file = client.file(file_id)
    with open(dest_path, "wb") as f:
        box_file.download_to(f)
    print(f"  Downloaded template from Box → {dest_path}")


def upload_to_box(local_path: str, folder_id: str, filename: str):
    client = get_box_client()
    folder = client.folder(folder_id)

    # If file already exists in folder, update it; otherwise upload new
    existing = {item.name: item for item in folder.get_items()}
    if filename in existing:
        existing[filename].update_contents(local_path)
        print(f"  Updated existing file in Box: {filename}")
    else:
        folder.upload(local_path, filename)
        print(f"  Uploaded new file to Box: {filename}")


# ============================================================
# STEP 3: Populate the HFC inputs file
# ============================================================

def clean_label(s):
    if not s:
        return ""
    return re.sub(r"<[^>]+>", "", str(s)).strip()


# SurveyCTO surveys frequently use these as missing/refused/don't-know codes.
# Clauses comparing the field to one of these are excluded from bound parsing.
MISSING_CODES = {-888, -999, -666, -777, -88, -99}


def parse_constraint_bounds(expr):
    """Extract (hard_min, hard_max) from a SurveyCTO constraint expression.

    Handles common patterns like:
        .>0 and .<100              -> (0, 100)
        .>=0 and .<10              -> (0, 10)
        .>= 0 or .= -888 or .= -999 -> (0, None)
        . between 0 and 24          -> (0, 24)

    Branches comparing the field to a missing-value code are dropped.
    Clauses referencing other variables (${var}) are dropped — those are
    cross-checks, not magnitude bounds. Returns (None, None) if nothing
    parseable was found.
    """
    if not expr:
        return (None, None)
    s = str(expr).strip().strip("()").strip()
    if not s:
        return (None, None)

    m = re.match(r"^\s*\.\s*between\s+(-?\d+(?:\.\d+)?)\s+and\s+(-?\d+(?:\.\d+)?)\s*$",
                 s, flags=re.IGNORECASE)
    if m:
        return (_to_int_if_whole(float(m.group(1))),
                _to_int_if_whole(float(m.group(2))))

    branches = re.split(r"\s+or\s+", s, flags=re.IGNORECASE)
    real_branches = []
    for br in branches:
        br_clean = br.strip().strip("()").strip()
        mc = re.match(r"^\s*\.\s*=\s*(-?\d+(?:\.\d+)?)\s*$", br_clean)
        if mc and float(mc.group(1)) in MISSING_CODES:
            continue
        if "${" in br_clean:
            continue
        real_branches.append(br_clean)

    if not real_branches:
        return (None, None)

    branch = real_branches[0]
    mins, maxs = [], []
    for cl in re.split(r"\s+and\s+", branch, flags=re.IGNORECASE):
        cl = cl.strip().strip("()").strip()
        if not cl or "${" in cl:
            continue
        m = re.match(r"^\s*\.\s*(>=|<=|>|<|=)\s*(-?\d+(?:\.\d+)?)\s*$", cl)
        if not m:
            continue
        op, val = m.group(1), float(m.group(2))
        if op == "=":
            continue
        if op in (">=", ">"):
            mins.append(val)
        else:
            maxs.append(val)

    hard_min = max(mins) if mins else None
    hard_max = min(maxs) if maxs else None
    return (_to_int_if_whole(hard_min), _to_int_if_whole(hard_max))


def _to_int_if_whole(v):
    if v is None:
        return None
    return int(v) if v == int(v) else v


def build_label_lookup(survey_path: str) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(survey_path, read_only=True, data_only=True)
    ws = wb["survey"]
    labels = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = row[1] if len(row) > 1 else None
        label = row[2] if len(row) > 2 else None
        if name:
            labels[str(name)] = clean_label(label)
    wb.close()
    return labels


def classify_survey_variables(survey_path: str):
    from openpyxl import load_workbook
    wb = load_workbook(survey_path, read_only=True, data_only=True)
    ws = wb["survey"]

    SKIP_TYPES = {
        "begin group", "end group", "begin repeat", "end repeat",
        "note", "calculate", "start", "end", "deviceid",
        "phonenumber", "username", "caseid", "enumerator",
    }

    groups = []
    numerics, selects, texts = [], [], []
    repeat_depth = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        def g(idx):
            return row[idx] if len(row) > idx and row[idx] else ""
        t, n, l = g(0), g(1), clean_label(g(2))
        constraint = g(9)
        disabled = str(g(12)).strip().lower() == "yes"
        t_str = str(t).lower() if t else ""

        # Track repeat-group nesting depth (skip disabled begin/end repeat
        # so depth stays balanced for them too)
        if "begin" in t_str and "repeat" in t_str:
            if not disabled:
                repeat_depth += 1
            continue
        if "end" in t_str and "repeat" in t_str:
            if not disabled:
                repeat_depth = max(0, repeat_depth - 1)
            continue

        if disabled:
            continue
        if not n:
            continue

        in_repeat = repeat_depth > 0
        base = t.split()[0] if t else ""
        if base in ("begin", "end") and "group" in t:
            if "begin" in t:
                groups.append({"name": n, "label": l, "in_repeat": in_repeat})
        elif base in SKIP_TYPES:
            continue
        elif base in ("integer", "decimal"):
            numerics.append({"name": n, "type": t, "label": l,
                             "in_repeat": in_repeat, "constraint": constraint})
        elif base in ("select_one", "select_multiple"):
            selects.append({"name": n, "type": t, "label": l, "in_repeat": in_repeat})
        elif base == "text":
            texts.append({"name": n, "label": l, "in_repeat": in_repeat})

    wb.close()
    return numerics, selects, texts, groups


def fmt_name(v):
    """Append * to variable name when it lives inside a repeat group, so
    Stata can match the wide-format instances (e.g. var_r1, var_r2 -> var_r*)."""
    name = str(v["name"]).strip()
    return f"{name}*" if v.get("in_repeat") else name


def existing_vars(ws, col=0):
    vals = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col]:
            vals.add(str(row[col]).strip())
    return vals


_OTHER_SUFFIXES = ("_oth", "_other", "_specify", "_oth_r", "_other_r", "_specify_r")


def populate_other_specify(ws, selects, texts):
    text_by_name = {v["name"]: v for v in texts}
    existing = existing_vars(ws, col=0)

    added = 0
    for v in selects:
        name = v["name"]
        for suffix in _OTHER_SUFFIXES:
            child = name + suffix
            child_var = text_by_name.get(child)
            if child_var and name not in existing:
                child_label = child_var.get("label") or "Other, specify"
                ws.append((fmt_name(v), v["label"], fmt_name(child_var), child_label,
                           None, None, "id member_name village enumerator_id"))
                existing.add(name)
                added += 1
                break
    return added


def populate_outliers(ws, numerics):
    existing = existing_vars(ws, col=0)
    added = 0
    for v in numerics:
        if v["name"] not in existing:
            combine = "yes" if v.get("in_repeat") else None
            ws.append((fmt_name(v), v["label"], "enum", "sd", 3, combine, None, None, "id member_name enumerator_id"))
            added += 1
    return added


def populate_constraints(ws, numerics):
    """Hard min/max are pulled directly from each numeric's SurveyCTO
    `constraint` expression. Soft bounds are intentionally left blank for
    the user to calibrate per study — there's no defensible way to invent
    them from the form alone."""
    existing = existing_vars(ws, col=0)
    added = 0
    for v in numerics:
        if v["name"] in existing:
            continue
        hard_min, hard_max = parse_constraint_bounds(v.get("constraint"))
        ws.append((fmt_name(v), v["label"], hard_min, None, None, hard_max, None, None, None))
        added += 1
    return added


def populate_logic(ws, numerics):
    """Logic checks are study-specific cross-variable consistency rules and
    cannot reliably be derived from XLSForm metadata alone — the constraint
    column rarely encodes them, and patterns like annual/monthly pairs are
    naming-convention-dependent. Sheet is left for the user to fill in."""
    return 0


def populate_enumstats(ws, numerics):
    existing = existing_vars(ws, col=0)
    added = 0
    for v in numerics:
        if v["name"] not in existing:
            combine = "yes" if v.get("in_repeat") else None
            ws.append((fmt_name(v), v["label"], "yes", None, "number", "yes", "yes", "yes", combine, None))
            added += 1
    return added


def populate_text_audits(ws, groups):
    """Emit every top-level begin-group from the form. The user can prune
    the ones they don't want timed in text audits."""
    existing = existing_vars(ws, col=0)
    added = 0
    for g in groups:
        if g["name"] not in existing:
            ws.append((fmt_name(g), None, None, g["label"]))
            added += 1
    return added


def populate_hfc_inputs(survey_path: str, template_path: str, output_path: str):
    from openpyxl import load_workbook

    print("\n  Classifying survey variables...")
    numerics, selects, texts, groups = classify_survey_variables(survey_path)
    print(f"    Found {len(numerics)} numeric, {len(selects)} select, {len(texts)} text variables")

    print("  Loading HFC template...")
    wb = load_workbook(template_path, keep_vba=True)

    results = {}
    results["other specify"] = populate_other_specify(wb["other specify"], selects, texts)
    results["outliers"]      = populate_outliers(wb["outliers"], numerics)
    results["constraints"]   = populate_constraints(wb["constraints"], numerics)
    results["logic"]         = populate_logic(wb["logic"], numerics)
    results["enumstats"]     = populate_enumstats(wb["enumstats"], numerics)
    results["text audits"]   = populate_text_audits(wb["text audits"], groups)

    wb.save(output_path)
    print(f"\n  Populated file saved → {output_path}")

    for sheet, count in results.items():
        print(f"    {sheet:20} +{count} rows added")

    return results


# ============================================================
# MAIN — orchestrates everything
# ============================================================

def main():
    print("=" * 60)
    print("HFC Inputs Auto-Populator")
    print("=" * 60)

    with tempfile.TemporaryDirectory() as tmpdir:
        survey_path   = os.path.join(tmpdir, "survey.xlsx")
        template_path = os.path.join(tmpdir, "hfc_template.xlsm")
        output_path   = os.path.join(tmpdir, OUTPUT_FILENAME)

        print("\nStep 1: Downloading survey from Google Drive...")
        download_from_google_drive(GDRIVE_SURVEY_FILE_ID, survey_path)

        print("\nStep 2: Downloading HFC template from Box...")
        download_from_box(BOX_TEMPLATE_FILE_ID, template_path)

        print("\nStep 3: Populating HFC inputs...")
        populate_hfc_inputs(survey_path, template_path, output_path)

        print("\nStep 4: Uploading populated file to Box...")
        upload_to_box(output_path, BOX_OUTPUT_FOLDER_ID, OUTPUT_FILENAME)

    print("\nDone! Your populated HFC inputs file is in Box.")
    print("=" * 60)


if __name__ == "__main__":
    # First time Box setup: uncomment the line below, run once, then comment it out again
    # box_first_time_auth(); exit()

    main()
